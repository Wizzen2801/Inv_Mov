import pandas as pd
from pandas import DataFrame
import sys
import numpy as np
from openpyxl import load_workbook

input_file = 'C:/Python_SHL/Test/OUT1.xlsx'
df1 = pd.read_excel(input_file)
df2 = pd.read_excel('C:/Python_SHL/Test/Code_Master.xlsx')

# df1 의 Product_Code 와 df2의 Product_Code 가 같으면 df1 에 Product_Name 을 추가하여 df2 의  Product_Name 을 매칭함
df = df1.join(df2.set_index('Product_Code')['Master_Price'], on='Product_Code')
output_file = 'C:/Python_SHL/Test/OUT_Temp.xlsx'
df.to_excel(output_file, index = 0)

wb = load_workbook(output_file)
ws = wb.active

for r in ws.rows:
    row_index = r[0].row
    ws.cell(row=row_index, column=4).value = "=B"+str(row_index)+"*C"+str(row_index)+''
    ws['D1'] ='OUT_Amt'
   
output_final = 'C:/Python_SHL/Test/OUT_Final.xlsx'
wb.save(output_final)
