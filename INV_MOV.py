import pandas as pd
from pandas import DataFrame
import sys
import numpy as np
from openpyxl import load_workbook

# 아래와 같은 컬럼명이 있는 파일 생성
df = DataFrame(columns = ['Product_Code', 'Product_Name', 'BOH_Qty', 'BOH_Amt', 'IN_Qty', 'IN_Amt', 'OUT_Qty', 'OUT_Amt', 'Sales_Gift_Qty', 'Sales_Gift_Amt', 'Sample_Qty', 'Sample_Amt', 'Warranty_Qty', 'Warranty_Amt', 'EOH_Qty', 'EOH_Amt'])
df.to_excel('C:/Python_WZN/Test/INV_MOV.xlsx', index=False)

# 아래 파일들을 모두 합쳐서 INV_MOV_Total.xlsx 로 생성
excel_names = ['c:/Python_WZN/Test/INV_MOV.xlsx', 'c:/Python_SHL/Test/BOH.xlsx', 'c:/Python_SHL/Test/IN.xlsx', 'c:/Python_SHL/Test/OUT.xlsx', 'c:/Python_SHL/Test/Sales_Gift.xlsx', 'c:/Python_SHL/Test/Sample.xlsx', 'c:/Python_SHL/Test/Warranty.xlsx', 'c:/Python_SHL/Test/EOH.xlsx']

excels = [pd.ExcelFile(name) for name in excel_names] 

frames = [x.parse(x.sheet_names[0], header=None,index_col=None) for x in excels]  
frames[1:] = [df[1:] for df in frames[1:]]  
combined = pd.concat(frames)

combined.to_excel('C:/Python_WZN/Test/INV_MOV_Total.xlsx', header=False, index=False)

# 생성된 파일을 Pivot 테이블로 변환
df = pd.read_excel('C:/Python_WZN/Test/INV_MOV_Total.xlsx')

df = df.pivot_table(index = ['Product_Code'], aggfunc = {'BOH_Qty':sum, 'BOH_Amt':sum, 'IN_Qty':sum, 'IN_Amt':sum, 'OUT_Qty':sum, 'OUT_Amt':sum, 'Sales_Gift_Qty':sum, 'Sales_Gift_Amt':sum, 'Sample_Qty':sum, 'Sample_Amt':sum, 'Warranty_Qty':sum, 'Warranty_Amt':sum, 'EOH_Qty':sum, 'EOH_Amt':sum})

# 제품명을 코드 마스터 테이블에서 가져와 매칭함
df2 = pd.read_excel('C:/Python_WZN/Test/Code_Master.xlsx')
df = df.join(df2.set_index('Product_Code')['Product_Name'], on='Product_Code')


# 피벗테이블로 변환시 칼럼 위치가 변경된 것을 다시 순서를 바로잡음
df = df.reindex(columns=['Product_Name', 'BOH_Qty','BOH_Amt','IN_Qty','IN_Amt', 'OUT_Qty', 'OUT_Amt', 'Sales_Gift_Qty', 'Sales_Gift_Amt', 'Sample_Qty', 'Sample_Amt', 'Warranty_Qty', 'Warranty_Amt', 'EOH_Qty', 'EOH_Amt'])
df.to_excel('C:/Python_WZN/Test/INV_MOV_Pivot.xlsx')

wb = load_workbook('C:/Python_WZN/Test/INV_MOV_Pivot.xlsx')
ws = wb.active

# Adjust_Qty, Amt 항목에 수식을 넣음
for r in ws.rows:
    row_index = r[0].row
    ws.cell(row=row_index, column=17).value = "=C"+str(row_index)+"+E"+str(row_index)+"-G"+str(row_index)+"-I"+str(row_index)+"-K"+str(row_index)+"-M"+str(row_index)+"-O"+str(row_index)+''
    ws['Q1'] ='Adjust_Qty'
    ws.cell(row=row_index, column=18).value = "=D"+str(row_index)+"+F"+str(row_index)+"-H"+str(row_index)+"-J"+str(row_index)+"-L"+str(row_index)+"-N"+str(row_index)+"-P"+str(row_index)+''
    ws['R1'] ='Adjust_Amt'
   
    wb.save('C:/Python_WZN/Test/INV_MOV_Final.xlsx')
